from django.shortcuts import render
from rest_framework import viewsets, filters, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import *
from .serializers import *
from .permissions import IsOwnerOrAdmin
from django.contrib.auth.models import Group
from rest_framework.generics import ListAPIView
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Count
from decimal import ROUND_HALF_UP, Decimal
from django.utils.timezone import now
import pandas as pd

#Generer fichier ABP128
from openpyxl import load_workbook
from io import BytesIO
from django.http import FileResponse

class BaseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    def perform_create(self, serializer):
        serializer.save(
            user_created=self.request.user,
            user_updated=self.request.user
        )

    def perform_update(self, serializer):
        serializer.save(user_updated=self.request.user)

class UserViewSet(BaseViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filterset_fields = ['nom', 'prenom', 'email', 'statut']
    search_fields = ['nom', 'prenom', 'email']
    ordering_fields = ['nom', 'created_at']

    # Ajoutez cette méthode à votre classe UserViewSet existante
    @action(detail=False, methods=['post'], permission_classes=[AllowAny])
    def register(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response(
                {"message": "Utilisateur créé avec succès"},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def connection_history(self, request, pk=None):
        user = self.get_object()
        data = {
            'username': user.username,
            'derniere_connexion': user.last_login.strftime("%Y-%m-%d %H:%M:%S") if user.last_login else None,
            'derniere_deconnexion': user.last_logout.strftime("%Y-%m-%d %H:%M:%S") if user.last_logout else None,
            'statut_actuel': 'Connecté' if user.is_active else 'Déconnecté'
        }
        return Response(data)
    
class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    
    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Group.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} roles supprimés avec succès",
            "deleted_count": deleted_count
        })

class AgenceViewSet(BaseViewSet):
    queryset = Agence.objects.all()
    serializer_class = AgenceSerializer
    filterset_fields = ['nom_agence']
    search_fields = ['nom_agence']
    ordering_fields = ['nom_agence', 'created_at']

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Agence.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} agence supprimées avec succès",
            "deleted_count": deleted_count
        })

class AllAgenceNopginListAPIView(ListAPIView):
    queryset = Agence.objects.all()
    serializer_class = AgenceSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class EtatViewSet(BaseViewSet):
    queryset = Etat.objects.all()
    serializer_class = EtatSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Etat.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })

class StatutViewSet(BaseViewSet):
    queryset = Statut.objects.all()
    serializer_class = StatutSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Statut.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} statut supprimées avec succès",
            "deleted_count": deleted_count
        })

class PublicStatutListAPIView(ListAPIView):
    queryset = Statut.objects.all()
    serializer_class = StatutSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class ServiceViewSet(BaseViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Service.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} departement supprimées avec succès",
            "deleted_count": deleted_count
        })

class TypeBesoinViewSet(BaseViewSet):
    queryset = TypeBesoin.objects.all()
    serializer_class = TypeBesoinSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = TypeBesoin.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} type besoin supprimées avec succès",
            "deleted_count": deleted_count
        })

class PrioriteViewSet(BaseViewSet):
    queryset = Priorite.objects.all()
    serializer_class = PrioriteSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Priorite.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} priorités supprimées avec succès",
            "deleted_count": deleted_count
        })

class UserServiceViewSet(BaseViewSet):
    queryset = UserService.objects.all()
    serializer_class = UserServiceSerializer

class UserAgenceViewSet(BaseViewSet):
    queryset = UserAgence.objects.all()
    serializer_class = UserAgenceSerializer

class BesoinViewSet(BaseViewSet):
    queryset = Besoin.objects.all()
    serializer_class = BesoinSerializer
    filterset_fields = ['reference', 'titre', 'typebesoin', 'etat', 'priorite', 'user']
    search_fields = ['reference', 'titre', 'description']
    ordering_fields = ['date_debut', 'date_fin', 'created_at']

    @action(detail=False, methods=['get'])
    def mes_besoins(self, request):
        besoins = self.queryset.filter(user=request.user)
        page = self.paginate_queryset(besoins)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)
    
    # ⚙️ Statistiques par état
    @action(detail=False, methods=['get'], url_path='statistiques-par-etat')
    def statistiques_par_etat(self, request):
        data = (
            self.queryset
            .values('etat__libelle_etat')
            .annotate(nombre=Count('id'))
            .order_by('etat__libelle_etat')
        )
        result = [{'etat': d['etat__libelle_etat'], 'nombre': d['nombre']} for d in data]
        return Response(result)

    # ⚙️ Statistiques par priorité
    @action(detail=False, methods=['get'], url_path='statistiques-par-priorite')
    def statistiques_par_priorite(self, request):
        data = (
            self.queryset
            .values('priorite__libelle')
            .annotate(nombre=Count('id'))
            .order_by('priorite__libelle')
        )
        result = [{'priorite': d['priorite__libelle'], 'nombre': d['nombre']} for d in data]
        return Response(result)
        # Dans BesoinViewSet

    @action(detail=False, methods=['get'], url_path='statistiques-globales')
    def statistiques_globales(self, request):
        # 1. Récupération brute
        etats = (
            self.queryset
            .values('etat__libelle_etat')
            .annotate(nombre=Count('id'))
            .order_by('etat__libelle_etat')
        )
        etat_result = [{'etat': d['etat__libelle_etat'], 'nombre': d['nombre']} for d in etats]

        priorites = (
            self.queryset
            .values('priorite__libelle')
            .annotate(nombre=Count('id'))
            .order_by('priorite__libelle')
        )
        priorite_result = [{'priorite': d['priorite__libelle'], 'nombre': d['nombre']} for d in priorites]

        # 2. Calcul du total général
        total_besoins = sum(item['nombre'] for item in etat_result)

        # 3. Filtrage des "clôturé(s)" et "annulé(s)"
        total_cloture = sum(item['nombre'] for item in etat_result if item['etat'].lower() in ['clôturé', 'cloturé', 'clôturés'])
        total_annule = sum(item['nombre'] for item in etat_result if item['etat'].lower() in ['annulé', 'annulés'])

        # 4. Calcul des non clôturés
        total_non_cloture = total_besoins - total_cloture - total_annule

        return Response({
            'par_etat': etat_result,
            'par_priorite': priorite_result,
            'total_non_cloture': total_non_cloture,
            'total_besoin': total_besoins
        })
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)

        if not serializer.is_valid():
            print("⚠️ Erreurs de validation dans le serializer:")
            print(serializer.errors)  # <-- Ce que l'on veut voir
            return Response(serializer.errors, status=400)

        self.perform_update(serializer)
        return Response(serializer.data)

class BesoinServiceViewSet(BaseViewSet):
    queryset = BesoinService.objects.all()
    serializer_class = BesoinServiceSerializer

class BesoinDocumentViewSet(BaseViewSet):
    queryset = BesoinDocument.objects.all()
    serializer_class = BesoinDocumentSerializer

class PublicTypeBesoinListAPIView(ListAPIView):
    queryset = TypeBesoin.objects.all()
    serializer_class = TypeBesoinSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class PublicPrioriteListAPIView(ListAPIView):
    queryset = Priorite.objects.all()
    serializer_class = PrioriteSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class BesoinCreateView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, format=None):
        serializer = BesoinWithDocumentsSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            besoin = serializer.save()  # 👈 on ne passe rien ici
            return Response(
                {
                    "message": "Besoin et documents enregistrés avec succès",
                    "besoin_id": besoin.id
                },
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



#Cantine

class TypePlatViewSet(BaseViewSet):
    queryset = TypePlat.objects.all()
    serializer_class = TypePlatSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = TypePlat.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })


class PublicTypePlatListAPIView(ListAPIView):
    queryset = TypePlat.objects.all()
    serializer_class = TypePlatSerializer
    permission_classes = [AllowAny]
    pagination_class = None
    

class PlatViewSet(BaseViewSet):
    queryset = Plat.objects.all()
    serializer_class = PlatSerializer
    filterset_fields = ['type_plat']
    search_fields = ['nom', 'description']
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def partial_update(self, request, *args, **kwargs):
        plat = self.get_object()
        data = request.data.copy()

        # 🔻 Supprimer les images cochées
        images_to_delete = request.data.getlist('images_to_delete[]', [])
        for img_id in images_to_delete:
            try:
                PlatImage.objects.get(id=img_id, plat=plat).delete()
            except PlatImage.DoesNotExist:
                pass  # image non trouvée = on ignore

        # ✅ Ajouter de nouvelles images
        new_images = request.FILES.getlist('new_images') 
        for image in new_images:
            PlatImage.objects.create(plat=plat, image=image)

        # 🔄 Mise à jour partielle des champs
        serializer = self.get_serializer(plat, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response({'message': 'Plat mis à jour avec succès'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Plat.objects.filter(id__in=ids).delete()
        return Response({
            "message": f"{deleted_count} plats supprimés avec succès",
            "deleted_count": deleted_count
        }, status=status.HTTP_200_OK)

class PlatImageViewSet(BaseViewSet):
    queryset = PlatImage.objects.all()
    serializer_class = PlatImageSerializer


class PlatCreateView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = PlatCreateWithImagesSerializer(data=request.data)
        if serializer.is_valid():
            # ✅ Ajout explicite des utilisateurs
            plat = serializer.save(
                user_created=request.user,
                user_updated=request.user
            )
            return Response({'message': 'Plat enregistré', 'id': plat.id}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    

class MenuViewSet(BaseViewSet):
    queryset = Menu.objects.all()
    serializer_class = MenuSerializer
    filterset_fields = ['date_menu']

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Menu.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })

class MenuPlatViewSet(BaseViewSet):
    queryset = MenuPlat.objects.all()
    serializer_class = MenuPlatSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = MenuPlat.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })
    

class PlatUploadAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        # images[] doit être une liste de fichiers
        fichiers = request.FILES.getlist('images[]')
        data = request.data.copy()

        # Fusionner les fichiers dans le format attendu par DRF (JSON)
        data.setlist('images', [])
        for fichier in fichiers:
            data.appendlist('images', {'image': fichier, 'is_principale': False})

        serializer = PlatUploadSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            plat = serializer.save()
            return Response({'message': 'Plat créé avec succès', 'plat_id': plat.id}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class CommandeViewSet(BaseViewSet):
    queryset = Commande.objects.all()
    serializer_class = CommandeSerializer
    filterset_fields = ['user', 'menu', 'plat']

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Commande.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })

class RetraitViewSet(BaseViewSet):
    queryset = Retrait.objects.all()
    serializer_class = RetraitSerializer
    filterset_fields = ['commande__user__matricule']

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Retrait.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} état supprimées avec succès",
            "deleted_count": deleted_count
        })


# Create your views here.

# Generation fichier APB128

class GenererAPB128_1View(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [AllowAny]

    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "Fichier manquant."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=excel_file, data_only=True)
            ws_ident = wb["Identification"]
            ws_prel = wb["Prelevements"]

            # Lecture des données d'identification
            ident_data = {
                row[0].value: str(row[1].value).strip() if row[1].value is not None else ''
                for row in ws_ident.iter_rows(min_row=2)
            }

            # Lecture des prélèvements
            prelevements = []
            for row in ws_prel.iter_rows(min_row=2, values_only=True):
                prelevements.append({
                    "Date_Execution": row[0],
                    "Code_Guichet": row[1],
                    "Compte": row[2],
                    "Cle_RIB": row[3],
                    "Nom_Client": row[4],
                    "Libelle_Banque_Client": row[5],
                    "Libelle_Operation": row[6],
                    "Numero_Autorisation": row[7],
                    "Montant": row[8]
                })

            # Fonctions utilitaires
            def left_pad(val, size, char='0'):
                return str(val).strip()[:size].rjust(size, char)

            def right_pad(val, size, char=' '):
                return str(val).strip()[:size].ljust(size, char)

            def format_amount(val):
                """Formatte le montant en entier sur 12 chiffres (centimes)."""
                val = int(val or 0)
                return str(val).rjust(12, '0')[-12:]

            lines = []
            seq = 1
            compte_total = 0
            montant_total = 0

            # Ligne 1 : Identification
            lines.append(
                "03" + "1" + left_pad(seq, 5) +
                left_pad(ident_data["Date_Execution"], 6) +
                "CI006" +
                left_pad(ident_data["Code_Guichet"], 5) +
                left_pad(ident_data["Compte"], 12) +
                left_pad(ident_data["Cle_RIB"], 2) +
                right_pad(ident_data["Nom_Emetteur"], 24) +
                left_pad(ident_data["Code_Emetteur"], 5) +
                right_pad(" ", 1) +
                left_pad(ident_data["Date_Remise"], 6) +
                right_pad(ident_data["Reference_Remise"], 7) +
                right_pad(ident_data["Zone_Libre"], 47)
            )

            # Lignes de prélèvement
            for p in prelevements:
                seq += 1
                compte_val = int(str(p["Compte"]).lstrip("0") or 0)
                montant_val = int(p["Montant"] or 0)

                compte_total += compte_val
                montant_total += montant_val

                lines.append(
                    "03" + "2" + left_pad(seq, 5) +
                    left_pad(p["Date_Execution"], 6) +
                    "CI006" +
                    left_pad(p["Code_Guichet"], 5) +
                    left_pad(p["Compte"], 12) +
                    left_pad(p["Cle_RIB"], 2) +
                    right_pad(p["Nom_Client"], 24) +
                    right_pad(p["Libelle_Banque_Client"], 17) +
                    right_pad(p["Libelle_Operation"], 20) +
                    left_pad(p["Numero_Autorisation"] or '', 10) +
                    format_amount(p["Montant"]) +
                    "00" + right_pad("", 5)
                )

            # Ligne de contrôle (type 9)
            seq += 1
            lines.append(
                "03" + "9" + left_pad(seq, 5) +
                left_pad(ident_data["Date_Remise"], 6) +
                right_pad("", 8) +
                format_amount(compte_total) +
                right_pad(ident_data["Compte"].lstrip("0"), 77) +
                format_amount(montant_total) +
                right_pad("", 5)
            )

            # Génération du fichier texte
            output = BytesIO()
            output.write("\n".join(lines).encode('utf-8'))
            output.seek(0)

            return FileResponse(output, as_attachment=True, filename="FICHIER_APB128.txt")

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class GenererAPB128View(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [AllowAny]

    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "Fichier manquant."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=excel_file, data_only=True)
            ws_ident = wb["Identification"]
            ws_prel = wb["Prelevements"]

            ident_data = {
                row[0].value: str(row[1].value).strip() if row[1].value is not None else ''
                for row in ws_ident.iter_rows(min_row=2)
            }

            prelevements = []
            for row in ws_prel.iter_rows(min_row=2, values_only=True):
                prelevements.append({
                    "Date_Execution": row[0],
                    "Code_Guichet": row[1],
                    "Compte": row[2],
                    "Cle_RIB": row[3],
                    "Nom_Client": row[4],
                    "Libelle_Banque_Client": row[5],
                    "Libelle_Operation": row[6],
                    "Numero_Autorisation": row[7],
                    "Montant": row[8]
                })

            def left_pad(val, size, char='0'):
                return str(val).strip()[:size].rjust(size, char)

            def right_pad(val, size, char=' '):
                return str(val).strip()[:size].ljust(size, char)

            def format_amount(val):
                val = int(val or 0)
                return str(val).rjust(12, '0')[-12:]

            lines = []
            seq = 1
            montant_total = 0

            compte_total = int(ident_data["Compte"]) if ident_data["Compte"].isdigit() else 0

            # Ligne 1 : Identification
            lines.append(
                "03" + "1" + left_pad(seq, 5) +
                left_pad(ident_data["Date_Execution"], 6) +
                "CI006" +
                left_pad(ident_data["Code_Guichet"], 5) +
                left_pad(ident_data["Compte"], 12) +
                left_pad(ident_data["Cle_RIB"], 2) +
                right_pad(ident_data["Nom_Emetteur"], 24) +
                left_pad(ident_data["Code_Emetteur"], 5) +
                right_pad(" ", 1) +
                left_pad(ident_data["Date_Remise"], 6) +
                right_pad(ident_data["Reference_Remise"], 7) +
                right_pad(ident_data["Zone_Libre"], 47)
            )

            for p in prelevements:
                seq += 1
                compte_val = int(str(p["Compte"]).lstrip("0") or 0)
                montant_val = int(p["Montant"] or 0)

                compte_total += compte_val
                montant_total += montant_val

                lines.append(
                    "03" + "2" + left_pad(seq, 5) +
                    left_pad(p["Date_Execution"], 6) +
                    "CI006" +
                    left_pad(p["Code_Guichet"], 5) +
                    left_pad(p["Compte"], 12) +
                    left_pad(p["Cle_RIB"], 2) +
                    right_pad(p["Nom_Client"], 24) +
                    right_pad(p["Libelle_Banque_Client"], 17) +
                    right_pad(p["Libelle_Operation"], 20) +
                    left_pad(p["Numero_Autorisation"] or "0000000000", 10) +
                    format_amount(p["Montant"]) +
                    "00" +
                    right_pad("", 5)
                )

            # Ligne 9 : Contrôle
            seq += 1
            lines.append(
                "03" + "9" + left_pad(seq, 5) +
                left_pad(ident_data["Date_Remise"], 6) +
                right_pad("", 8) +
                format_amount(compte_total) +          # ✅ Cumul compte (émetteur + clients)
                right_pad(ident_data["Compte"].lstrip("0"), 77) +                # ✅ Zone libre de 77 pos remplie de zéros
                format_amount(montant_total) +          # ✅ Cumul montants
                right_pad("", 5)
            )

            output = BytesIO()
            output.write("\n".join(lines).encode('utf-8'))
            output.seek(0)

            return FileResponse(output, as_attachment=True, filename="FICHIER_APB128.txt")

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


# Debut Gestion entretien vehicule
class TypeVehiculeViewSet(BaseViewSet):
    queryset = TypeVehicule.objects.all()
    serializer_class = TypeVehiculeSerializer

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = TypeVehicule.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} type véhicule supprimés avec succès",
            "deleted_count": deleted_count
        })
    
class VehiculeViewSet(BaseViewSet):
    queryset = Vehicule.objects.all().order_by('id')
    serializer_class = VehiculeSerializer
    filterset_fields = ['immatriculation']

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Vehicule.objects.filter(id__in=ids).delete()
        
        return Response({
            "message": f"{deleted_count} véhicule supprimés avec succès",
            "deleted_count": deleted_count
        })

class PublicVehiculeListAPIView(ListAPIView):
    queryset = Vehicule.objects.all()
    serializer_class = VehiculeSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class PublicTypeVehiculeListAPIView(ListAPIView):
    queryset = TypeVehicule.objects.all()
    serializer_class = TypeVehiculeSerializer
    permission_classes = [AllowAny]
    pagination_class = None

class ImportEntretienExcelView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [AllowAny]

    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "Fichier manquant"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active

            lignes_traitees = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                immatricule_str, description, periode, debit, credit, solde = row

                if not immatricule_str or not periode:
                    continue

                try:
                    vehicule = Vehicule.objects.get(immatriculation=immatricule_str)
                except Vehicule.DoesNotExist:
                    continue

                EntretienVehicule.objects.filter(
                    vehicule=vehicule,
                    periode=periode
                ).delete()

                EntretienVehicule.objects.create(
                    vehicule=vehicule,
                    description=description or '',
                    date_entretien_vehicule=now(),
                    periode=periode,
                    debit=Decimal(debit or 0),
                    credit=Decimal(credit or 0),
                    solde=Decimal(solde or 0)
                )

                lignes_traitees += 1

            return Response({
                "success": f"{lignes_traitees} entretiens importés avec succès"
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EntretienVehiculeViewSet(BaseViewSet):
    queryset = EntretienVehicule.objects.all().order_by('periode')
    serializer_class = EntretienVehiculeSerializer
    filterset_fields = ['vehicule']  # renommé ici

    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])

        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "Veuillez fournir une liste d'IDs valides"}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = EntretienVehicule.objects.filter(id__in=ids).delete()

        return Response({
            "message": f"{deleted_count} entretien véhicule supprimés avec succès",
            "deleted_count": deleted_count
        })

class EtatEntretienGroupesView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        # 🔹 Récupération des paramètres
        annee = request.query_params.get('annee')
        agence_id = request.query_params.get('agence_id')
        immatriculation = request.query_params.get('immatriculation')

        # 🔸 Gestion de l'année
        if not annee:
            annee = datetime.today().year
        else:
            try:
                annee = int(annee)
            except ValueError:
                return Response({"error": "Année invalide"}, status=400)

        # 🔸 Gestion de l'agence (optionnelle)
        if agence_id:
            try:
                agence_id = int(agence_id)
            except ValueError:
                return Response({"error": "agence_id invalide"}, status=400)
        else:
            agence_id = None

        # 🔸 Appel de la fonction de regroupement
        data = get_entretien_par_agence_avec_groupement(
            annee=annee,
            agence_id=agence_id,
            immatriculation=immatriculation
        )

        return Response(data)
#Fin entretien vehicule


# Ressources Humaines
class SimulationSalaireUploadView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [AllowAny]

    def post(self, request):
        fichier = request.FILES.get("file")
        if not fichier:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(fichier)
            simulations = []

            for _, row in df.iterrows():
                # 🔍 Lecture sécurisée de la date
                try:
                    valeur_date = row["Date de naiss."]
                    if isinstance(valeur_date, (datetime, pd.Timestamp)):
                        date_naiss = valeur_date.date()
                    else:
                        date_naiss = datetime.strptime(str(valeur_date).strip(), "%d/%m/%Y").date()
                except Exception:
                    return Response({
                        "error": f"Date invalide pour le matricule {row['Matricule']}. La colonne 'Date de naiss.' doit contenir des dates valides."
                    }, status=status.HTTP_400_BAD_REQUEST)

                # 🔢 Lecture et calculs
                age_2025 = 2025 - date_naiss.year
                salaire = float(row["Salaire de base brut 2025"])

                simulation = SimulationSalaire.objects.create(
                    matricule=row["Matricule"],
                    categorie=row["Cat."],
                    date_naissance=date_naiss,
                    salaire_base=row["Salaire de base"],
                    sursalaire=row["Sursalaire"],
                    salaire_brut_2025=salaire
                )

                annee = 2025
                salaire_actuel = salaire

                while age_2025 <= 60:
                    if salaire_actuel < 150000:
                        taux = 0.20
                    elif salaire_actuel < 250000:
                        taux = 0.10
                    else:
                        taux = 0.025

                    augmentation = round(salaire_actuel * taux)
                    cotisation = round(salaire_actuel * 0.075) if salaire_actuel > 250000 else 0

                    SimulationHistorique.objects.create(
                        simulation=simulation,
                        annee=annee,
                        age=age_2025,
                        salaire=salaire_actuel,
                        taux_augmentation=taux * 100,
                        augmentation=augmentation,
                        cotisation=cotisation
                    )

                    salaire_actuel += augmentation
                    annee += 1
                    age_2025 += 1

                simulations.append(SimulationSalaireSerializer(simulation).data)

            return Response(simulations, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SimulationSalaireListView(ListAPIView):
    queryset = SimulationSalaire.objects.all().order_by('-created_at')
    serializer_class = SimulationSalaireSerializer


class SimulationViewSet(BaseViewSet):
    queryset = SimulationSalaire.objects.all().order_by('matricule')
    serializer_class = SimulationSalaireSerializer
    permission_classes = [AllowAny]
    pagination_class = None

    def get_serializer_context(self):
        return {'request': self.request}

    def list(self, request, *args, **kwargs):
        # Filtres
        annee_debut = request.query_params.get('annee_debut')
        annee_fin = request.query_params.get('annee_fin')
        salaire_min = request.query_params.get('salaire_min')
        salaire_max = request.query_params.get('salaire_max')

        filters = Q()
        if annee_debut:
            filters &= Q(annee__gte=annee_debut)
        if annee_fin:
            filters &= Q(annee__lte=annee_fin)
        if salaire_min:
            filters &= Q(salaire__gte=salaire_min)
        if salaire_max:
            filters &= Q(salaire__lte=salaire_max)

        queryset = self.filter_queryset(self.get_queryset().prefetch_related('historique'))

        filtered_queryset = []
        for sim in queryset:
            historique_qs = sim.historique.all()
            if filters:
                historique_qs = historique_qs.filter(filters)
            if historique_qs.exists():
                filtered_queryset.append(sim)

        count = len(filtered_queryset)

        page = self.paginate_queryset(filtered_queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            data = serializer.data
        else:
            serializer = self.get_serializer(filtered_queryset, many=True)
            data = serializer.data

        historiques = SimulationHistorique.objects.filter(filters) if filters else SimulationHistorique.objects.all()

        # Totaux globaux
        total_global = {
            "salaire": sum(float(h.salaire) for h in historiques),
            "augmentation": sum(float(h.augmentation) for h in historiques),
            "cotisation": sum(float(h.cotisation) for h in historiques),
        }

        # Totaux par année
        total_par_annee = {}
        for h in historiques:
            annee = str(h.annee)
            if annee not in total_par_annee:
                total_par_annee[annee] = {
                    "salaire": 0, "augmentation": 0, "cotisation": 0,
                    "effectif_aug": 0, "effectif_cot": 0
                }

            total_par_annee[annee]["salaire"] += float(h.salaire)
            total_par_annee[annee]["augmentation"] += float(h.augmentation)
            total_par_annee[annee]["cotisation"] += float(h.cotisation)

            if float(h.augmentation) > 0:
                total_par_annee[annee]["effectif_aug"] += 1
            if float(h.cotisation) > 0:
                total_par_annee[annee]["effectif_cot"] += 1

        # ✅ Ajouter les effectifs non augmenté / non cotisé
        for annee, totals in total_par_annee.items():
            totals["effectif_non_aug"] = count - totals["effectif_aug"]
            totals["effectif_non_cot"] = count - totals["effectif_cot"]

        # Totaux par année et tranche
        def init_tranche():
            return {
                "salaire": 0, "augmentation": 0, "cotisation": 0,
                "effectif_aug": 0, "effectif_cot": 0
            }

        total_par_annee_tranche_20 = {}
        total_par_annee_tranche_10 = {}
        total_par_annee_tranche_2_5 = {}

        for h in historiques:
            annee = str(h.annee)
            salaire = float(h.salaire)
            augmentation = float(h.augmentation)
            cotisation = float(h.cotisation)

            if salaire < 150000:
                target = total_par_annee_tranche_20
            elif salaire <= 250000:
                target = total_par_annee_tranche_10
            else:
                target = total_par_annee_tranche_2_5

            if annee not in target:
                target[annee] = init_tranche()

            target[annee]["salaire"] += salaire
            target[annee]["augmentation"] += augmentation
            target[annee]["cotisation"] += cotisation

            if augmentation > 0:
                target[annee]["effectif_aug"] += 1
            if cotisation > 0:
                target[annee]["effectif_cot"] += 1

        # Ajout des effectifs non augmentés / non cotisés dans chaque tranche
        for total in [total_par_annee_tranche_20, total_par_annee_tranche_10, total_par_annee_tranche_2_5]:
            for annee, vals in total.items():
                vals["effectif_non_aug"] = count - vals["effectif_aug"]
                vals["effectif_non_cot"] = count - vals["effectif_cot"]

        return Response({
            "results": data,
            "total_global": total_global,
            "total_par_annee": total_par_annee,
            "total_par_annee_tranche_20": total_par_annee_tranche_20,
            "total_par_annee_tranche_10": total_par_annee_tranche_10,
            "total_par_annee_tranche_2_5": total_par_annee_tranche_2_5,
            "count": count
        })

    # Generer simulation historique salaire
class GenererSimulationHistorique2AncienView(APIView):
    def post(self, request):
        # 🔁 Purge des données existantes
        deleted_count, _ = SimulationHistorique2.objects.all().delete()

        simulations = SimulationSalaire.objects.all()
        total_crees = 0

        for simulation in simulations:
            age_2025 = 2025 - simulation.date_naissance.year
            salaire = float(simulation.salaire_brut_2025)

            annee = 2025
            age = age_2025

            while age <= 60:
                # Déterminer le taux selon les tranches
                if salaire < 150000:
                    taux = 0.20
                elif salaire < 250000:
                    taux = 0.10
                else:
                    taux = 0.025

                augmentation = round(salaire * taux)
                cotisation = round(salaire * 0.075) if salaire > 250000 else 0

                # Création dans SimulationHistorique2
                SimulationHistorique2.objects.create(
                    simulation=simulation,
                    annee=annee,
                    age=age,
                    salaire=salaire,
                    taux_augmentation=taux * 100,
                    augmentation=augmentation,
                    cotisation=cotisation
                )

                salaire += augmentation
                annee += 1
                age += 1
                total_crees += 1

        return Response({
            "message": f"{deleted_count} anciennes lignes supprimées. {total_crees} nouvelles lignes créées dans SimulationHistorique2."
        }, status=status.HTTP_201_CREATED)
    

class GenererSimulationHistorique2View(APIView):
    def post(self, request):
        # 🔁 Purge des anciennes données
        deleted_count, _ = SimulationHistorique2.objects.all().delete()

        simulations = SimulationSalaire.objects.all()
        total_crees = 0

        for simulation in simulations:
            age_2025 = 2025 - simulation.date_naissance.year
            salaire = Decimal(simulation.salaire_brut_2025).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

            annee = 2025
            age = age_2025
            is_first_year = True

            while age <= 60:
                # 🎯 Logique spécifique pour la première année
                if is_first_year and salaire < 150000:
                    augmentation = Decimal(150000 - salaire).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                    taux = (Decimal(augmentation) / salaire * 100).quantize(Decimal('0.01'))
                    salaire = Decimal(150000)
                else:
                    if salaire < 150000:
                        taux = Decimal('20.00')
                    elif salaire <= 250000:
                        taux = Decimal('10.00')
                    else:
                        taux = Decimal('2.50')

                    augmentation = (salaire * taux / 100).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

                cotisation = (salaire * Decimal('0.075')).quantize(Decimal('1.')) if salaire > 250000 else Decimal(0)

                SimulationHistorique2.objects.create(
                    simulation=simulation,
                    annee=annee,
                    age=age,
                    salaire=salaire,
                    taux_augmentation=taux,
                    augmentation=augmentation,
                    cotisation=cotisation
                )

                salaire += augmentation
                annee += 1
                age += 1
                total_crees += 1
                is_first_year = False

        return Response({
            "message": f"{deleted_count} anciennes lignes supprimées. {total_crees} lignes créées dans SimulationHistorique2."
        }, status=status.HTTP_201_CREATED)


class GenererSimulationHistorique22222View(APIView):
    def post(self, request):
        # 🔁 Purge des anciennes données
        deleted_count, _ = SimulationHistorique2.objects.all().delete()

        simulations = SimulationSalaire.objects.all()
        total_crees = 0

        for simulation in simulations:
            age_2025 = 2025 - simulation.date_naissance.year
            salaire = Decimal(simulation.salaire_brut_2025).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

            annee = 2025
            age = age_2025

            while age <= 60:
                # ✅ En 2025, on ne modifie pas le salaire
                if annee == 2025:
                    taux = Decimal('0.00')
                    augmentation = Decimal('0.00')
                else:
                    # 🎯 À partir de 2026, appliquer ta règle
                    if salaire < 150000:
                        augmentation = Decimal(150000 - salaire).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                        taux = (augmentation / salaire * 100).quantize(Decimal('0.01'))
                        salaire = Decimal('150000')
                    elif salaire <= 250000:
                        taux = Decimal('10.00')
                        augmentation = (salaire * taux / 100).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                    else:
                        taux = Decimal('2.50')
                        augmentation = (salaire * taux / 100).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

                # 🧮 Cotisation uniquement si > 250000
                cotisation = (salaire * Decimal('0.075')).quantize(Decimal('1.')) if salaire > 250000 else Decimal(0)

                # ✅ Enregistrement dans SimulationHistorique2
                SimulationHistorique2.objects.create(
                    simulation=simulation,
                    annee=annee,
                    age=age,
                    salaire=salaire,
                    taux_augmentation=taux,
                    augmentation=augmentation,
                    cotisation=cotisation
                )

                salaire += augmentation
                annee += 1
                age += 1
                total_crees += 1

        return Response({
            "message": f"{deleted_count} anciennes lignes supprimées. {total_crees} lignes créées dans SimulationHistorique2."
        }, status=status.HTTP_201_CREATED)
    

class Simulation2ViewSet(BaseViewSet):
    queryset = SimulationSalaire.objects.all().order_by('matricule')
    serializer_class = SimulationSalaire2Serializer
    permission_classes = [AllowAny]
    pagination_class = None

    def get_serializer_context(self):
        return {'request': self.request}

    def list(self, request, *args, **kwargs):
        # Filtres
        annee_debut = request.query_params.get('annee_debut')
        annee_fin = request.query_params.get('annee_fin')
        salaire_min = request.query_params.get('salaire_min')
        salaire_max = request.query_params.get('salaire_max')

        filters = Q()
        if annee_debut:
            filters &= Q(annee__gte=annee_debut)
        if annee_fin:
            filters &= Q(annee__lte=annee_fin)
        if salaire_min:
            filters &= Q(salaire__gte=salaire_min)
        if salaire_max:
            filters &= Q(salaire__lte=salaire_max)

        queryset = self.filter_queryset(self.get_queryset().prefetch_related('historique2'))

        filtered_queryset = []
        for sim in queryset:
            historique_qs = sim.historique2.all()
            if filters:
                historique_qs = historique_qs.filter(filters)
            if historique_qs.exists():
                filtered_queryset.append(sim)

        count = len(filtered_queryset)

        page = self.paginate_queryset(filtered_queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            data = serializer.data
        else:
            serializer = self.get_serializer(filtered_queryset, many=True)
            data = serializer.data

        historiques = SimulationHistorique2.objects.filter(filters) if filters else SimulationHistorique2.objects.all()

        # Totaux globaux
        total_global = {
            "salaire": sum(float(h.salaire) for h in historiques),
            "augmentation": sum(float(h.augmentation) for h in historiques),
            "cotisation": sum(float(h.cotisation) for h in historiques),
        }

        # Totaux par année
        total_par_annee = {}
        for h in historiques:
            annee = str(h.annee)
            if annee not in total_par_annee:
                total_par_annee[annee] = {
                    "salaire": 0, "augmentation": 0, "cotisation": 0,
                    "effectif_aug": 0, "effectif_cot": 0
                }

            total_par_annee[annee]["salaire"] += float(h.salaire)
            total_par_annee[annee]["augmentation"] += float(h.augmentation)
            total_par_annee[annee]["cotisation"] += float(h.cotisation)

            if float(h.augmentation) > 0:
                total_par_annee[annee]["effectif_aug"] += 1
            if float(h.cotisation) > 0:
                total_par_annee[annee]["effectif_cot"] += 1

        # ✅ Ajouter les effectifs non augmenté / non cotisé
        for annee, totals in total_par_annee.items():
            totals["effectif_non_aug"] = count - totals["effectif_aug"]
            totals["effectif_non_cot"] = count - totals["effectif_cot"]

        # Totaux par année et tranche
        def init_tranche():
            return {
                "salaire": 0, "augmentation": 0, "cotisation": 0,
                "effectif_aug": 0, "effectif_cot": 0
            }

        total_par_annee_tranche_20 = {}
        total_par_annee_tranche_10 = {}
        total_par_annee_tranche_2_5 = {}

        for h in historiques:
            annee = str(h.annee)
            salaire = float(h.salaire)
            augmentation = float(h.augmentation)
            cotisation = float(h.cotisation)

            if salaire < 150000:
                target = total_par_annee_tranche_20
            elif salaire <= 250000:
                target = total_par_annee_tranche_10
            else:
                target = total_par_annee_tranche_2_5

            if annee not in target:
                target[annee] = init_tranche()

            target[annee]["salaire"] += salaire
            target[annee]["augmentation"] += augmentation
            target[annee]["cotisation"] += cotisation

            if augmentation > 0:
                target[annee]["effectif_aug"] += 1
            if cotisation > 0:
                target[annee]["effectif_cot"] += 1

        # Ajout des effectifs non augmentés / non cotisés dans chaque tranche
        for total in [total_par_annee_tranche_20, total_par_annee_tranche_10, total_par_annee_tranche_2_5]:
            for annee, vals in total.items():
                vals["effectif_non_aug"] = count - vals["effectif_aug"]
                vals["effectif_non_cot"] = count - vals["effectif_cot"]

        return Response({
            "results": data,
            "total_global": total_global,
            "total_par_annee": total_par_annee,
            "total_par_annee_tranche_20": total_par_annee_tranche_20,
            "total_par_annee_tranche_10": total_par_annee_tranche_10,
            "total_par_annee_tranche_2_5": total_par_annee_tranche_2_5,
            "count": count
        })


# Fin ressources Humaines